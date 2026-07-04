"""Unit tests for the Excel payroll parser and validations."""

import io
import unittest
from decimal import Decimal
import openpyxl

from utils.excel_parser import parse_month_year, clean_phone, parse_date_str, parse_payroll_excel


class TestExcelParser(unittest.TestCase):
    """Verifies parsing of month strings, phone cleaning, and spreadsheet row parsing."""

    def test_parse_month_year(self):
        self.assertEqual(parse_month_year("06/2026"), ("June", 2026))
        self.assertEqual(parse_month_year("June 2026"), ("June", 2026))
        self.assertEqual(parse_month_year("June-2026"), ("June", 2026))
        self.assertIsNone(parse_month_year("invalid_month"))

    def test_clean_phone(self):
        self.assertEqual(clean_phone("9876543210"), "+919876543210")
        self.assertEqual(clean_phone("+919876543210"), "+919876543210")
        self.assertEqual(clean_phone("+1 123 456 7890"), "+11234567890")

    def test_parse_date_str(self):
        self.assertEqual(parse_date_str("2026-06-29"), "29/06/2026")
        self.assertEqual(parse_date_str("29/06/2026"), "29/06/2026")

    def test_parse_payroll_excel_valid(self):
        # Create a mock Excel spreadsheet in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = [
            "phone", "month_year", "establishment", "principal_employer", "address",
            "employee_name", "workman_id", "guardian_name", "designation", "uan",
            "bank_account", "wage_period", "attendance", "basic", "da", "allowances",
            "gross_wages", "pf", "esi", "other_deductions", "net_wages", "issue_date"
        ]
        ws.append(headers)
        
        row_data = [
            "9876543210", "June 2026", "Establishment 1", "Employer 1", "Address 1",
            "John Doe", "EMP001", "Father Doe", "Worker", "UAN12345",
            "BANK12345", "June-2026", "26.0", "10000", "2000", "1000",
            "13000", "1000", "500", "500", "11000", "29/06/2026"
        ]
        ws.append(row_data)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        
        valid, invalid = parse_payroll_excel(excel_bytes)
        
        self.assertEqual(len(valid), 1)
        self.assertEqual(len(invalid), 0)
        self.assertEqual(valid[0]["workman_id"], "EMP001")
        self.assertEqual(valid[0]["employee_name"], "John Doe")
        self.assertEqual(valid[0]["phone"], "+919876543210")
        self.assertEqual(valid[0]["net_wages"], 11000.0)

    def test_parse_payroll_excel_invalid_math_accepted(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = [
            "phone", "month_year", "establishment", "principal_employer", "address",
            "employee_name", "workman_id", "guardian_name", "designation", "uan",
            "bank_account", "wage_period", "attendance", "basic", "da", "allowances",
            "gross_wages", "pf", "esi", "other_deductions", "net_wages", "issue_date"
        ]
        ws.append(headers)
        
        # Net wages mismatch: Gross 13000 - ded 2000 = 11000 (not 9000)
        # This should now be parsed successfully and placed in the valid list.
        row_data = [
            "9876543210", "June 2026", "Establishment 1", "Employer 1", "Address 1",
            "John Doe", "EMP001", "Father Doe", "Worker", "UAN12345",
            "BANK12345", "June-2026", "26.0", "10000", "2000", "1000",
            "13000", "1000", "500", "500", "9000", "29/06/2026"
        ]
        ws.append(row_data)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        
        valid, invalid = parse_payroll_excel(excel_bytes)
        
        self.assertEqual(len(valid), 1)
        self.assertEqual(len(invalid), 0)
        self.assertEqual(valid[0]["net_wages"], 9000.0)

    def test_boundary_and_leap_years(self):
        # 1. Check boundary year/month parsing
        self.assertEqual(parse_month_year("12/2026"), ("December", 2026))
        self.assertEqual(parse_month_year("01/2027"), ("January", 2027))
        
        # 2. Check leap year date parsing (February 2028 is a leap year)
        self.assertEqual(parse_date_str("29/02/2028"), "29/02/2028")
        
        # 3. Check invalid leap year date parsing (February 2029 is not a leap year)
        # Note: parse_date_str falls back to returning the original string on parse failure
        self.assertEqual(parse_date_str("29/02/2029"), "29/02/2029")

    def test_duplicate_row_validation(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = [
            "phone", "month_year", "establishment", "principal_employer", "address",
            "employee_name", "workman_id", "guardian_name", "designation", "uan",
            "bank_account", "wage_period", "attendance", "basic", "da", "allowances",
            "gross_wages", "pf", "esi", "other_deductions", "net_wages", "issue_date"
        ]
        ws.append(headers)
        
        # Employee 1 - First occurrence
        ws.append([
            "9876543210", "June 2026", "Establishment 1", "Employer 1", "Address 1",
            "John Doe", "EMP001", "Father Doe", "Worker", "UAN12345",
            "BANK12345", "June-2026", "26.0", "10000", "2000", "1000",
            "13000", "1000", "500", "500", "11000", "29/06/2026"
        ])
        
        # Employee 1 - Duplicate occurrence in the same sheet for the same period
        ws.append([
            "9876543210", "June 2026", "Establishment 1", "Employer 1", "Address 1",
            "John Doe", "EMP001", "Father Doe", "Worker", "UAN12345",
            "BANK12345", "June-2026", "26.0", "10000", "2000", "1000",
            "13000", "1000", "500", "500", "11000", "29/06/2026"
        ])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        
        valid, invalid = parse_payroll_excel(excel_bytes)
        
        # First row should be valid, second row should fail validation as a duplicate
        self.assertEqual(len(valid), 1)
        self.assertEqual(len(invalid), 1)
        self.assertEqual(valid[0]["workman_id"], "EMP001")
        self.assertEqual(invalid[0]["data"]["workman_id"], "EMP001")
        self.assertIn("Duplicate employee entry", invalid[0]["error"])


if __name__ == "__main__":
    unittest.main()
